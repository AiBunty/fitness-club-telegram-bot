"""
Excel bulk upload for store products
Handles bulk creation, updates, and price revisions
"""

import logging
import io
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, Document
from telegram.ext import (
    ContextTypes,
    ConversationHandler,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)

from src.database.store_items_operations import add_or_update_item_in_db
from src.utils.auth import is_admin_id

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel upload disabled")

# Conversation states
EXCEL_UPLOAD = 1


async def cmd_store_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /store_excel command - Upload Excel for bulk product management
    """
    if not is_admin_id(update.effective_user.id):
        await update.message.reply_text("❌ Admin access only.")
        return ConversationHandler.END
    
    if not OPENPYXL_AVAILABLE:
        await update.message.reply_text(
            "❌ Excel support not installed.\n"
            "Install: pip install openpyxl"
        )
        return ConversationHandler.END
    
    text = (
        "📊 *Store Items Excel Upload*\n\n"
        "Send an Excel file with these columns:\n"
        "1. item_name\n"
        "2. hsn\n"
        "3. mrp\n"
        "4. gst_percent\n\n"
        "Rules:\n"
        "• Item Name + HSN is unique key\n"
        "• If name+hsn exists → UPDATE\n"
        "• If new → CREATE\n"
        "• Upload is idempotent\n\n"
        "📥 Upload your Excel file:"
    )
    
    keyboard = [[InlineKeyboardButton("📥 Download Sample", callback_data="excel_sample")]]
    
    await update.message.reply_text(
        text,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    
    return EXCEL_UPLOAD


async def handle_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Process uploaded Excel file"""
    
    if not update.message.document:
        await update.message.reply_text("❌ Please upload a file")
        return EXCEL_UPLOAD
    
    document = update.message.document
    
    # Check file type
    if not document.file_name.endswith(('.xlsx', '.xls', '.csv')):
        await update.message.reply_text("❌ Please upload Excel file (.xlsx, .xls, .csv)")
        return EXCEL_UPLOAD
    
    admin_id = update.effective_user.id
    
    try:
        # Download file
        file = await context.bot.get_file(document.file_id)
        file_bytes = io.BytesIO()
        await file.download_to_memory(file_bytes)
        file_bytes.seek(0)
        
        # Parse Excel
        workbook = openpyxl.load_workbook(file_bytes)
        sheet = workbook.active
        
        # Expected columns
        expected_columns = {
            1: 'item_name',
            2: 'hsn',
            3: 'mrp',
            4: 'gst_percent'
        }
        
        # Read header
        header_row = sheet[1]
        header_map = {}
        for col_num, col_name in expected_columns.items():
            cell = header_row[col_num - 1]
            if cell.value and cell.value.lower().strip() in [col_name, col_name.replace('_', ' '), col_name.replace('_', '-')]:
                header_map[col_num] = col_name
        
        if len(header_map) < len(expected_columns):
            await update.message.reply_text(
                "❌ Excel header mismatch.\n"
                f"Expected columns: {', '.join(expected_columns.values())}"
            )
            return EXCEL_UPLOAD
        
        # Process rows
        success_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            try:
                item_name = row[0].value
                hsn = row[1].value
                mrp = row[2].value
                gst_percent = row[3].value
                
                # Validate required fields
                if not item_name or not hsn or mrp is None:
                    errors.append(f"Row {row_num}: Missing required fields")
                    error_count += 1
                    continue
                
                # Type conversions
                try:
                    mrp = float(mrp)
                    gst_percent = float(gst_percent or 18)
                except ValueError as e:
                    errors.append(f"Row {row_num}: Invalid data type - {str(e)}")
                    error_count += 1
                    continue
                
                # Create/update product
                result = add_or_update_item_in_db({
                    'name': str(item_name).strip(),
                    'hsn': str(hsn).strip(),
                    'mrp': mrp,
                    'gst': gst_percent
                })
                
                if result:
                    success_count += 1
                else:
                    errors.append(f"Row {row_num}: Failed to upsert product")
                    error_count += 1
            
            except Exception as e:
                logger.error(f"Error processing row {row_num}: {e}")
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        # Send summary
        summary = (
            f"📊 *Excel Upload Summary*\n\n"
            f"✅ Success: {success_count}\n"
            f"❌ Errors: {error_count}\n"
        )
        
        if errors and len(errors) <= 10:
            summary += f"\n📋 Errors:\n" + "\n".join(errors[:10])
        elif errors:
            summary += f"\n📋 First 10 errors:\n" + "\n".join(errors[:10])
            summary += f"\n... and {len(errors) - 10} more errors"
        
        await update.message.reply_text(summary, parse_mode="Markdown")
        
        logger.info(f"Excel upload by admin {admin_id}: {success_count} success, {error_count} errors")
        
        return ConversationHandler.END
    
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        await update.message.reply_text(f"❌ Error processing file: {str(e)}")
        return EXCEL_UPLOAD


async def excel_sample(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Download sample Excel template"""
    query = update.callback_query
    await query.answer()
    
    try:
        # Create sample workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        
        # Header
        headers = ['item_name', 'hsn', 'mrp', 'gst_percent']
        sheet.append(headers)
        
        # Sample data
        samples = [
            ['Formula 1 Shake', '2106', 1500, 18],
            ['Aloe Concentrate', '2202', 1750, 18],
            ['Afresh Energy Drink', '2106', 750, 18],
        ]
        
        for sample in samples:
            sheet.append(sample)
        
        # Save to bytes
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Send file
        await query.edit_message_text("📥 Sending sample Excel template...")
        await context.bot.send_document(
            chat_id=query.from_user.id,
            document=excel_file,
            filename="store_items_sample.xlsx",
            caption="📊 Sample template for bulk item upload"
        )
        
    except Exception as e:
        logger.error(f"Error creating sample: {e}")
        await query.edit_message_text(f"❌ Error: {str(e)}")


def get_store_excel_conversation_handler():
    """Return Excel upload conversation handler"""
    return ConversationHandler(
        entry_points=[
            CommandHandler("store_excel", cmd_store_excel),
        ],
        states={
            EXCEL_UPLOAD: [
                MessageHandler(filters.Document.ALL, handle_excel_file),
                CallbackQueryHandler(excel_sample, pattern="^excel_sample$"),
            ],
        },
        fallbacks=[]
    )
