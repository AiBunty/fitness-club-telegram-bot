import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import ContextTypes, ConversationHandler, MessageHandler, filters
from typing import Dict
from io import BytesIO
from src.utils.gst import load_gst_config, save_gst_config, is_gst_enabled, get_gst_mode, get_gst_percent
from src.utils.store_items import add_or_update_item, load_store_items, find_items
from src.utils.state_management import clear_stale_states
from src.utils.auth import is_admin_id

logger = logging.getLogger(__name__)

# GST settings conversation states
GST_TOGGLE, GST_MODE, GST_PERCENT = range(3)

# Store item conversation states
ITEM_NAME, ITEM_HSN, ITEM_MRP, ITEM_GST, BULK_UPLOAD_AWAIT, DELETE_SERIAL_OR_NAME, EDIT_SERIAL_OR_NAME, EDIT_FIELD, EDIT_VALUE = range(3, 12)


async def cmd_gst_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Admin-only: verify role
    if not is_admin_id(update.effective_user.id):
        if update.callback_query:
            await update.callback_query.answer("❌ Admin access only.", show_alert=True)
        else:
            await update.message.reply_text("❌ Admin access only.")
        return ConversationHandler.END
    query = update.callback_query
    if query:
        await query.answer()
        clear_stale_states(update, context, flow_name="gst_settings")
        message = query.message
    else:
        message = update.message

    cfg = load_gst_config()
    enabled = cfg.get('enabled', False)
    mode = cfg.get('mode', 'exclusive')
    percent = cfg.get('percent', 0)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Toggle GST ON", callback_data="gst_toggle_on") if not enabled else InlineKeyboardButton("Toggle GST OFF", callback_data="gst_toggle_off")],
        [InlineKeyboardButton("Mode: Inclusive" if mode=='inclusive' else "Mode: Exclusive", callback_data="gst_change_mode")],
        [InlineKeyboardButton("Edit Percentage", callback_data="gst_edit_percent")],
        [InlineKeyboardButton("⬅ Back", callback_data="admin_dashboard")]
    ])

    text = f"🧾 GST Settings\n\nEnabled: {enabled}\nMode: {mode}\nPercent: {percent}%"
    await message.reply_text(text, reply_markup=kb)
    return ConversationHandler.END


async def gst_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_id(update.effective_user.id):
        await update.callback_query.answer("❌ Admin access only.", show_alert=True)
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    cfg = load_gst_config()
    if query.data == 'gst_toggle_on':
        cfg['enabled'] = True
    else:
        cfg['enabled'] = False
    save_gst_config(cfg)
    await query.edit_message_text(f"✅ GST set to {cfg['enabled']}.")
    return ConversationHandler.END


async def gst_change_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_id(update.effective_user.id):
        await update.callback_query.answer("❌ Admin access only.", show_alert=True)
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    cfg = load_gst_config()
    cfg['mode'] = 'inclusive' if cfg.get('mode') == 'exclusive' else 'exclusive'
    save_gst_config(cfg)
    await query.edit_message_text(f"✅ GST mode changed to {cfg['mode']}")
    return ConversationHandler.END


async def gst_edit_percent_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_id(update.effective_user.id):
        await update.callback_query.answer("❌ Admin access only.", show_alert=True)
        return ConversationHandler.END
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Enter GST percentage (0-100):")
    return GST_PERCENT


async def gst_edit_percent(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_id(update.effective_user.id):
        await update.message.reply_text("❌ Admin access only.")
        return ConversationHandler.END
    text = update.message.text.strip()
    try:
        p = float(text)
        if p < 0 or p > 100:
            await update.message.reply_text("❌ GST must be between 0 and 100. Try again:")
            return GST_PERCENT
        cfg = load_gst_config()
        cfg['percent'] = p
        save_gst_config(cfg)
        await update.message.reply_text(f"✅ GST percent set to {p}%")
        return ConversationHandler.END
    except Exception:
        await update.message.reply_text("❌ Enter a valid number (e.g., 18). Try again:")
        return GST_PERCENT


# Store item flows
async def cmd_create_store_items(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Admin-only: verify role
    if not is_admin_id(update.effective_user.id):
        if update.callback_query:
            await update.callback_query.answer("❌ Admin access only.", show_alert=True)
        else:
            await update.message.reply_text("❌ Admin access only.")
        return ConversationHandler.END
    query = update.callback_query
    if query:
        await query.answer()
        clear_stale_states(update, context, flow_name="store_admin")
        message = query.message
    else:
        message = update.message

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Create Item", callback_data="store_create_item")],
        [InlineKeyboardButton("📥 Bulk Upload (Excel)", callback_data="store_bulk_upload")],
        [InlineKeyboardButton("📄 Download Sample Excel", callback_data="store_download_sample")],
        [InlineKeyboardButton("📄 Download Existing Items", callback_data="store_download_existing")],
        [InlineKeyboardButton("🗑 Delete Item", callback_data="store_delete_item")],
        [InlineKeyboardButton("✏️ Edit Item Price/GST", callback_data="store_edit_item")],
        [InlineKeyboardButton("⬅ Back", callback_data="admin_dashboard")],
    ])
    await message.reply_text("🏬 Store Items Master\n\nChoose action:", reply_markup=kb)
    return ConversationHandler.END


async def store_create_item_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    clear_stale_states(update, context, flow_name="store_admin")
    logger.info("[STORE_CREATE] entering create item flow")
    await query.edit_message_text("Enter Item Name:")
    return ITEM_NAME


async def store_item_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    logger.info(f"[STORE_CREATE] state=ITEM_NAME input={name}")
    if not name:
        await update.message.reply_text("❌ Name cannot be empty. Try again:")
        return ITEM_NAME
    context.user_data['store_item'] = {'name': name}
    await update.message.reply_text("Enter HSN Code:")
    return ITEM_HSN


async def store_item_hsn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    hsn = update.message.text.strip()
    context.user_data['store_item']['hsn'] = hsn
    await update.message.reply_text("Enter MRP:")
    return ITEM_MRP


async def store_item_mrp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        mrp = float(update.message.text.strip())
        if mrp <= 0:
            await update.message.reply_text("❌ MRP must be > 0. Try again:")
            return ITEM_MRP
        context.user_data['store_item']['mrp'] = mrp
        # default GST from global
        from src.utils.gst import get_gst_percent
        context.user_data['store_item']['gst'] = get_gst_percent()
        await update.message.reply_text(f"Enter GST % for item (default {context.user_data['store_item']['gst']}):")
        return ITEM_GST
    except Exception:
        await update.message.reply_text("❌ Enter a valid MRP (e.g., 499.00). Try again:")
        return ITEM_MRP


async def store_item_gst(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        g = float(update.message.text.strip())
        if g < 0 or g > 100:
            await update.message.reply_text("❌ GST must be 0–100. Try again:")
            return ITEM_GST
        item = context.user_data.pop('store_item')
        item['gst'] = g
        
        # Add item with serial number
        result = add_or_update_item({
            'name': item['name'],
            'hsn': item.get('hsn',''),
            'mrp': item['mrp'],
            'gst': item['gst']
        })
        
        # Get serial from result
        serial = result.get('serial', '?')
        logger.info(f"[STORE_CREATE] item_saved serial={serial} name={item['name']}")
        
        await update.message.reply_text(
            f"✅ Item created successfully\n"
            f"Serial: {serial}\n"
            f"Name: {item['name']}"
        )
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"[STORE_CREATE] error saving item: {e}")
        await update.message.reply_text("❌ Enter a valid GST (e.g., 18). Try again:")
        return ITEM_GST


# Bulk upload: send sample excel and accept file upload
async def store_bulk_upload_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    clear_stale_states(update, context, flow_name="store_admin")
    logger.info("[STORE_BULK] entering bulk upload flow")
    # generate sample excel (with Serial No column present)
    try:
        import asyncio
        from openpyxl import Workbook
        
        # Wrap Excel generation in thread to prevent blocking
        def generate_sample_excel():
            wb = Workbook()
            ws = wb.active
            # Header with Serial No included
            ws.append(['Serial No','Item Name','HSN Code','MRP','GST %'])
            # Example rows: one empty serial (new), one with serial for update
            ws.append(['', 'Sample New Item', '1001', 499.00, 18])
            ws.append([1, 'Formula q', '4819', 1800.00, 18])
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio
        
        # Run in separate thread to avoid blocking
        bio = await asyncio.to_thread(generate_sample_excel)
        
        await query.message.reply_document(document=InputFile(bio, filename='store_items_sample.xlsx'))
        await query.message.reply_text('Upload filled Excel file (as attachment). Ensure Serial No column present — leave empty for NEW items.')
        return BULK_UPLOAD_AWAIT  # Stay in conversation to await document
    except Exception as e:
        logger.error(f"[STORE_BULK] failed to send sample Excel: {e}")
        await query.message.reply_text('Could not generate sample Excel. Ensure openpyxl is installed.')
        return ConversationHandler.END


async def handle_uploaded_store_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # User uploaded a file (document)
    doc = update.message.document
    if not doc:
        await update.message.reply_text('No file found. Upload the Excel file.')
        return BULK_UPLOAD_AWAIT
    
    logger.info(f"[STORE_BULK] upload_received filename={doc.file_name}")
    
    # Send immediate feedback to admin
    await update.message.reply_text('⏳ Processing file, please wait...')
    
    file = await doc.get_file()
    bio = BytesIO()
    await file.download_to_memory(out=bio)
    bio.seek(0)
    
    try:
        import asyncio
        from openpyxl import load_workbook
        
        # Wrap Excel parsing in thread to prevent blocking
        def parse_excel(bio):
            wb = load_workbook(filename=bio, data_only=True)
            ws = wb.active
            return list(ws.iter_rows(values_only=True))
        
        # Run in separate thread
        rows = await asyncio.to_thread(parse_excel, bio)
        
        if len(rows) < 2:
            await update.message.reply_text('❌ Excel file is empty or has no data rows.')
            return ConversationHandler.END
        
        # Expected header includes Serial No
        header = [str(c).strip().lower() if c else '' for c in rows[0]]
        # Allow flexible column ordering but require these headers
        try:
            serial_i = header.index('serial no')
            name_i = header.index('item name')
            hsn_i = header.index('hsn code')
            mrp_i = header.index('mrp')
            gst_i = header.index('gst %')
        except ValueError:
            await update.message.reply_text('❌ Excel header mismatch. Required columns: Serial No, Item Name, HSN Code, MRP, GST %')
            return ConversationHandler.END

        rows_read = 0
        new_items = 0
        updated_items = 0
        skipped = 0
        serial_conflicts = []

        for r in rows[1:]:
            rows_read += 1
            try:
                # Read values safely
                serial_cell = r[serial_i] if serial_i < len(r) else None
                serial_val = serial_cell if serial_cell is None else serial_cell
                serial_raw = str(serial_val).strip() if serial_val not in (None, '') else ''

                name = str(r[name_i]).strip() if name_i < len(r) and r[name_i] is not None else ''
                hsn = str(r[hsn_i]).strip() if hsn_i < len(r) and r[hsn_i] is not None else ''
                mrp_raw = r[mrp_i] if mrp_i < len(r) else None
                gst_raw = r[gst_i] if gst_i < len(r) else None

                # Validation
                if not name or not hsn:
                    skipped += 1
                    continue
                try:
                    mrp = float(mrp_raw)
                    gst = float(gst_raw)
                except Exception:
                    skipped += 1
                    continue
                if mrp <= 0 or gst < 0 or gst > 100:
                    skipped += 1
                    continue

                # Case A: Serial present
                if serial_raw:
                    try:
                        s = int(float(serial_raw))
                    except Exception:
                        skipped += 1
                        continue
                    # Update by serial only
                    try:
                        res = add_or_update_item({'serial': s, 'name': name, 'hsn': hsn, 'mrp': mrp, 'gst': gst})
                        updated_items += 1
                    except KeyError:
                        # serial not found -> skip
                        skipped += 1
                        serial_conflicts.append(serial_raw)
                        continue
                else:
                    # No serial: check by name+hsn
                    res = add_or_update_item({'name': name, 'hsn': hsn, 'mrp': mrp, 'gst': gst})
                    if res.get('is_new'):
                        new_items += 1
                    else:
                        updated_items += 1

            except Exception as e:
                logger.warning(f"[STORE_BULK] skipped row due to exception: {e}")
                skipped += 1
                continue
        
        logger.info(f"[STORE_EXCEL] rows_read={rows_read}")
        logger.info(f"[STORE_EXCEL] new_items={new_items}")
        logger.info(f"[STORE_EXCEL] updated_items={updated_items}")
        logger.info(f"[STORE_EXCEL] skipped_rows={skipped}")
        if serial_conflicts:
            logger.info(f"[STORE_EXCEL] serial_conflict={serial_conflicts}")

        await update.message.reply_text(
            f"✅ Bulk upload completed\n\n"
            f"➕ New items added: {new_items}\n"
            f"✏ Items updated (price/GST): {updated_items}\n"
            f"⚠ Rows skipped: {skipped}"
        )
        return ConversationHandler.END
        
    except Exception as e:
        logger.error(f"[STORE_BULK] error parsing excel: {e}")
        await update.message.reply_text('❌ Failed to parse Excel. Ensure it is a valid .xlsx file with required columns.\n\nPlease upload a valid Excel file or type /cancel to exit.')
        return BULK_UPLOAD_AWAIT


async def store_download_existing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Generate an Excel of existing store items and send to admin."""
    query = update.callback_query
    if query:
        await query.answer()
        message = query.message
    else:
        message = update.message

    try:
        import asyncio
        from openpyxl import Workbook

        def generate_excel():
            items = load_store_items()
            wb = Workbook()
            ws = wb.active
            ws.append(['Serial No', 'Item Name', 'HSN Code', 'MRP', 'GST %'])
            # sort by serial ascending
            try:
                items_sorted = sorted(items, key=lambda x: int(x.get('serial', 0)))
            except Exception:
                items_sorted = items
            for it in items_sorted:
                ws.append([
                    it.get('serial', ''),
                    it.get('name', ''),
                    it.get('hsn', ''),
                    float(it.get('mrp', 0)) if it.get('mrp', None) is not None else '',
                    float(it.get('gst', 0)) if it.get('gst', None) is not None else ''
                ])
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio

        bio = await asyncio.to_thread(generate_excel)
        await message.reply_document(document=InputFile(bio, filename='store_items_current.xlsx'))
        return ConversationHandler.END
    except Exception as e:
        logger.error(f"[STORE_DOWNLOAD] failed to generate/send excel: {e}")
        await message.reply_text('❌ Failed to generate store items Excel. Ensure openpyxl is installed.')
        return ConversationHandler.END


# Store item search API used by invoice flow (DEPRECATED - DO NOT USE)
# This function should NOT be registered as a handler
# It's kept only for legacy compatibility
async def search_store_items(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    DEPRECATED: This should not be called directly.
    Use Invoice v2 store search instead.
    """
    logger.warning("[STORE_SEARCH] DEPRECATED search_store_items called - should not happen!")
    term = update.message.text.strip()
    results = find_items(term, limit=10)
    if not results:
        await update.message.reply_text('No items found. Try another name.')
        return
    for it in results:
        text = f"{it.get('name')}\nHSN: {it.get('hsn','')}\nMRP: ₹{float(it.get('mrp',0)):.2f}\nGST: {it.get('gst',0)}%"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton('Select', callback_data=f"store_select_{it.get('name')}_{it.get('hsn')}")]])
        await update.message.reply_text(text, reply_markup=kb)


async def store_item_select_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # callback_data format: store_select_<name>_<hsn>
    parts = query.data.split('_', 2)
    if len(parts) < 3:
        await query.edit_message_text('Invalid selection')
        return
    payload = parts[2]
    # payload contains name_hsn combined; split last underscore
    # naive parse: try to find last underscore separating name and hsn
    try:
        raw = query.data[len('store_select_'):]
        # split from right
        name, hsn = raw.rsplit('_', 1)
    except Exception:
        name = raw
        hsn = ''
    # find item
    items = load_store_items()
    chosen = None
    for it in items:
        if it.get('name') == name and str(it.get('hsn','')) == hsn:
            chosen = it
            break
    if not chosen:
        await query.edit_message_text('❌ Item not found')
        return
    # Store selection in conversation context so invoice flow can pick it up
    context.user_data['invoice']['current_item'] = {
        'name': chosen.get('name'),
        'rate': float(chosen.get('mrp',0)),
        'gst': float(chosen.get('gst',0))
    }
    await query.edit_message_text(f"Item selected: {chosen.get('name')}\nRate: ₹{float(chosen.get('mrp',0)):.2f}")
    # Prompt quantity next (invoice flow expects quantity then discount)
    await context.bot.send_message(chat_id=query.from_user.id, text='🔢 Enter quantity:')
    return


async def handle_bulk_upload_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text messages during bulk upload (user sent text instead of file)"""
    logger.info(f"[STORE_BULK] received text instead of file: {update.message.text}")
    await update.message.reply_text(
        '❌ Please upload a valid Excel file (.xlsx) as a document attachment.\n\n'
        'Tap the 📎 (attachment) icon and select your Excel file.\n\n'
        'Or type /cancel to exit.'
    )
    return BULK_UPLOAD_AWAIT


# ==================== DELETE ITEM FLOW ====================
async def store_delete_item_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Prompt admin to enter serial or name to delete"""
    query = update.callback_query
    await query.answer()
    clear_stale_states(update, context, flow_name="store_admin")
    await query.edit_message_text(
        "🗑 Delete Item\n\n"
        "Enter Serial Number or Item Name to search:"
    )
    return DELETE_SERIAL_OR_NAME


async def handle_delete_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Search for item to delete"""
    search_term = update.message.text.strip()
    
    # Try serial first
    try:
        serial = int(search_term)
        from src.database.store_items_operations import get_item_by_serial_from_db
        item = get_item_by_serial_from_db(serial)
        if item:
            context.user_data['delete_item'] = item
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Confirm Delete", callback_data=f"confirm_delete_{item['serial']}")],
                [InlineKeyboardButton("❌ Cancel", callback_data="cmd_create_store_items")]
            ])
            await update.message.reply_text(
                f"Found item:\n\n"
                f"Serial: {item['serial']}\n"
                f"Name: {item['name']}\n"
                f"HSN: {item['hsn']}\n"
                f"MRP: Rs {item['mrp']}\n"
                f"GST: {item['gst']}%\n\n"
                f"⚠️ Confirm deletion?",
                reply_markup=kb
            )
            return ConversationHandler.END
    except ValueError:
        pass
    
    # Search by name
    from src.database.store_items_operations import search_items_by_name_from_db
    items = search_items_by_name_from_db(search_term)
    
    if not items:
        await update.message.reply_text("❌ No items found. Try again or /cancel to exit.")
        return DELETE_SERIAL_OR_NAME
    
    if len(items) == 1:
        item = items[0]
        context.user_data['delete_item'] = item
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Confirm Delete", callback_data=f"confirm_delete_{item['serial']}")],
            [InlineKeyboardButton("❌ Cancel", callback_data="cmd_create_store_items")]
        ])
        await update.message.reply_text(
            f"Found item:\n\n"
            f"Serial: {item['serial']}\n"
            f"Name: {item['name']}\n"
            f"HSN: {item['hsn']}\n"
            f"MRP: Rs {item['mrp']}\n"
            f"GST: {item['gst']}%\n\n"
            f"⚠️ Confirm deletion?",
            reply_markup=kb
        )
        return ConversationHandler.END
    
    # Multiple matches
    text = f"Found {len(items)} items:\n\n"
    for item in items[:10]:
        text += f"Serial {item['serial']}: {item['name']} (Rs {item['mrp']})\n"
    text += "\nEnter Serial Number to delete:"
    await update.message.reply_text(text)
    return DELETE_SERIAL_OR_NAME


async def confirm_delete_item(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Confirm and delete the item"""
    query = update.callback_query
    await query.answer()
    
    serial = int(query.data.split("_")[-1])
    
    try:
        from src.database.store_items_operations import delete_item_from_db
        success = delete_item_from_db(serial)
        if success:
            await query.edit_message_text(f"✅ Item #{serial} deleted successfully!")
        else:
            await query.edit_message_text(f"❌ Failed to delete item #{serial}")
    except Exception as e:
        logger.error(f"Error deleting item: {e}")
        await query.edit_message_text(f"❌ Error deleting item: {e}")
    
    return ConversationHandler.END


# ==================== EDIT ITEM FLOW ====================
async def store_edit_item_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Prompt admin to enter serial or name to edit"""
    query = update.callback_query
    await query.answer()
    clear_stale_states(update, context, flow_name="store_admin")
    await query.edit_message_text(
        "✏️ Edit Item Price/GST\n\n"
        "Enter Serial Number or Item Name to search:"
    )
    return EDIT_SERIAL_OR_NAME


async def handle_edit_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Search for item to edit"""
    search_term = update.message.text.strip()
    
    # Try serial first
    try:
        serial = int(search_term)
        from src.database.store_items_operations import get_item_by_serial_from_db
        item = get_item_by_serial_from_db(serial)
        if item:
            context.user_data['edit_item'] = item
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("💰 Edit MRP", callback_data="edit_field_mrp")],
                [InlineKeyboardButton("📊 Edit GST", callback_data="edit_field_gst")],
                [InlineKeyboardButton("❌ Cancel", callback_data="cmd_create_store_items")]
            ])
            await update.message.reply_text(
                f"Current details:\n\n"
                f"Serial: {item['serial']}\n"
                f"Name: {item['name']}\n"
                f"HSN: {item['hsn']}\n"
                f"MRP: Rs {item['mrp']}\n"
                f"GST: {item['gst']}%\n\n"
                f"What would you like to edit?",
                reply_markup=kb
            )
            return ConversationHandler.END
    except ValueError:
        pass
    
    # Search by name
    from src.database.store_items_operations import search_items_by_name_from_db
    items = search_items_by_name_from_db(search_term)
    
    if not items:
        await update.message.reply_text("❌ No items found. Try again or /cancel to exit.")
        return EDIT_SERIAL_OR_NAME
    
    if len(items) == 1:
        item = items[0]
        context.user_data['edit_item'] = item
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💰 Edit MRP", callback_data="edit_field_mrp")],
            [InlineKeyboardButton("📊 Edit GST", callback_data="edit_field_gst")],
            [InlineKeyboardButton("❌ Cancel", callback_data="cmd_create_store_items")]
        ])
        await update.message.reply_text(
            f"Current details:\n\n"
            f"Serial: {item['serial']}\n"
            f"Name: {item['name']}\n"
            f"HSN: {item['hsn']}\n"
            f"MRP: Rs {item['mrp']}\n"
            f"GST: {item['gst']}%\n\n"
            f"What would you like to edit?",
            reply_markup=kb
        )
        return ConversationHandler.END
    
    # Multiple matches
    text = f"Found {len(items)} items:\n\n"
    for item in items[:10]:
        text += f"Serial {item['serial']}: {item['name']} (Rs {item['mrp']})\n"
    text += "\nEnter Serial Number to edit:"
    await update.message.reply_text(text)
    return EDIT_SERIAL_OR_NAME


async def prompt_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Prompt for new value"""
    query = update.callback_query
    await query.answer()
    
    field = query.data.split("_")[-1]  # mrp or gst
    context.user_data['edit_field'] = field
    
    item = context.user_data.get('edit_item', {})
    current_value = item.get(field, 0)
    
    if field == 'mrp':
        await query.edit_message_text(f"Current MRP: Rs {current_value}\n\nEnter new MRP:")
    else:
        await query.edit_message_text(f"Current GST: {current_value}%\n\nEnter new GST percentage:")
    
    return EDIT_VALUE


async def handle_edit_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Save edited value"""
    try:
        new_value = float(update.message.text.strip())
        field = context.user_data.get('edit_field')
        item = context.user_data.get('edit_item')
        
        if not item or not field:
            await update.message.reply_text("❌ Session expired. Please start again.")
            return ConversationHandler.END
        
        # Validate
        if field == 'mrp' and new_value <= 0:
            await update.message.reply_text("❌ MRP must be greater than 0. Try again:")
            return EDIT_VALUE
        if field == 'gst' and (new_value < 0 or new_value > 100):
            await update.message.reply_text("❌ GST must be between 0 and 100. Try again:")
            return EDIT_VALUE
        
        # Update in database
        from src.database.store_items_operations import add_or_update_item_in_db
        item[field] = new_value
        result = add_or_update_item_in_db({
            'serial': item['serial'],
            'name': item['name'],
            'hsn': item['hsn'],
            'mrp': item['mrp'],
            'gst': item['gst']
        })
        
        if result:
            await update.message.reply_text(
                f"✅ Updated successfully!\n\n"
                f"Serial: {result['serial']}\n"
                f"Name: {result['name']}\n"
                f"MRP: Rs {result['mrp']}\n"
                f"GST: {result['gst']}%"
            )
        else:
            await update.message.reply_text("❌ Failed to update item")
        
    except ValueError:
        await update.message.reply_text("❌ Invalid number. Try again:")
        return EDIT_VALUE
    except Exception as e:
        logger.error(f"Error updating item: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    
    return ConversationHandler.END


def get_store_and_gst_handlers():
    from telegram.ext import CallbackQueryHandler
    gst_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cmd_gst_settings, pattern='^cmd_gst_settings$')],
        states={
            GST_PERCENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, gst_edit_percent)]
        },
        fallbacks=[],
        conversation_timeout=300,  # 5 minutes timeout to prevent stuck states
        per_message=False,
        per_chat=True,  # CRITICAL: Isolate per chat for 200+ users
        per_user=True   # CRITICAL: Isolate per user for admin concurrency
    )

    store_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(cmd_create_store_items, pattern='^cmd_create_store_items$'),
            CallbackQueryHandler(store_create_item_prompt, pattern='^store_create_item$'),
            CallbackQueryHandler(store_bulk_upload_prompt, pattern='^store_bulk_upload$'),
            CallbackQueryHandler(store_bulk_upload_prompt, pattern='^store_download_sample$'),
            CallbackQueryHandler(store_download_existing, pattern='^store_download_existing$'),
            CallbackQueryHandler(store_delete_item_prompt, pattern='^store_delete_item$'),
            CallbackQueryHandler(store_edit_item_prompt, pattern='^store_edit_item$'),
        ],
        states={
            ITEM_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, store_item_name)],
            ITEM_HSN: [MessageHandler(filters.TEXT & ~filters.COMMAND, store_item_hsn)],
            ITEM_MRP: [MessageHandler(filters.TEXT & ~filters.COMMAND, store_item_mrp)],
            ITEM_GST: [MessageHandler(filters.TEXT & ~filters.COMMAND, store_item_gst)],
            BULK_UPLOAD_AWAIT: [
                MessageHandler(filters.Document.ALL, handle_uploaded_store_excel),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_bulk_upload_text)
            ],
            DELETE_SERIAL_OR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_delete_search)],
            EDIT_SERIAL_OR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_search)],
            EDIT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_value)],
        },
        fallbacks=[],
        conversation_timeout=300,  # 5 minutes timeout to prevent stuck states
        per_message=False,
        per_chat=True,  # CRITICAL: Isolate per chat for 200+ users
        per_user=True   # CRITICAL: Isolate per user for admin concurrency
    )

    return gst_conv, store_conv