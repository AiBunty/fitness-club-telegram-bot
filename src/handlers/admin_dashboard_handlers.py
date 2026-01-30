"""
Enhanced Admin Dashboard with Member Management & Export
- Member list with filtering
- Excel export functionality  
- User management (delete/ban)
- Analytics and reports
"""

import logging
import io
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler, CommandHandler
from src.database.user_operations import (
    get_user,
    delete_user,
    ban_user,
    unban_user,
    is_user_banned,
    get_manage_users_page,
    get_manage_users_count,
    MANAGE_USERS_FILTERS
)
from src.utils.auth import is_admin
from src.utils.message_templates import get_template, save_template
from src.utils import event_registry
from src.utils.state_management import clear_stale_states
from src.utils.flow_manager import (
    set_active_flow, clear_active_flow, check_flow_ownership,
    FLOW_DELETE_USER, FLOW_BAN_USER
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

logger = logging.getLogger(__name__)

# Conversation states for user management (high-range to avoid collisions)
MANAGE_USER_MENU, SELECT_USER_ACTION, CONFIRM_DELETE, ENTER_BAN_REASON = range(5000, 5004)

# Manage Users list defaults
MANAGE_USERS_PER_PAGE = 25
MANAGE_USERS_INACTIVE_DAYS = 90


def _normalize_filter(filter_type: str) -> str:
    filter_type = (filter_type or 'all').lower()
    return filter_type if filter_type in MANAGE_USERS_FILTERS else 'all'


def _filter_label(filter_type: str) -> str:
    labels = {
        'all': 'All Users',
        'paid': 'Paid Users',
        'unpaid': 'Unpaid Users',
        'active': 'Active (Paid)',
        'inactive': f'Inactive {MANAGE_USERS_INACTIVE_DAYS}d (Unpaid)'
    }
    return labels.get(filter_type, 'All Users')


async def cmd_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Main admin panel with all options"""
    if not is_admin(update.effective_user.id):
        if update.callback_query:
            await update.callback_query.answer("❌ Admin access only.")
        else:
            await update.message.reply_text("❌ Admin access only.")
        return
    
    # CRITICAL: Clear any active conversation states to prevent cross-talk
    # This ensures that abandoned Invoice/Store/User searches don't interfere
    # with new admin actions (e.g., Store Item creation thinking it's User Search)
    if context.user_data:
        logger.info(f"[ADMIN_PANEL] Clearing active states: {list(context.user_data.keys())}")
        context.user_data.clear()
    
    # Handle both message and callback contexts
    if update.callback_query:
        await update.callback_query.answer()
        message = update.callback_query.message
    else:
        message = update.message
    
    keyboard = [
        [InlineKeyboardButton("👥 Member List", callback_data="admin_members_list"),
         InlineKeyboardButton("📊 Dashboard", callback_data="admin_dashboard")],
        [InlineKeyboardButton("👤 Manage Users", callback_data="admin_manage_users"),
         InlineKeyboardButton("📥 Excel Export", callback_data="admin_export_excel")],
        [InlineKeyboardButton("🧾 Invoices", callback_data="cmd_invoices"),
         InlineKeyboardButton("📋 Invoice Reports", callback_data="cmd_invoice_reports")],
        [InlineKeyboardButton("💰 Revenue Stats", callback_data="dashboard_revenue"),
         InlineKeyboardButton("📈 Engagement", callback_data="dashboard_engagement")],
        [InlineKeyboardButton("✏️ Message Templates", callback_data="admin_templates")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await message.reply_text(
        "🔧 *Admin Control Panel*\n\n"
        "Select an option below:",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


# --- Template management conversation ---
SELECT_TEMPLATE, ENTER_TEMPLATE_TEXT = range(2)


async def cmd_templates_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return
    await query.answer()

    # List event keys with preview/edit buttons
    keyboard = []
    for k in sorted(event_registry.EVENT_KEYS):
        keyboard.append([InlineKeyboardButton(f"Preview {k}", callback_data=f"template_preview:{k}"),
                         InlineKeyboardButton(f"Edit {k}", callback_data=f"template_edit:{k}")])

    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="admin_dashboard_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        text="✏️ *Message Templates*\n\nSelect an event to preview or edit:",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


async def cmd_followup_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return
    await query.answer()

    from src.utils.followup_manager import load_followups
    followups = load_followups()
    keyboard = []
    for k in sorted(event_registry.EVENT_KEYS):
        keyboard.append([InlineKeyboardButton(f"Edit follow-ups for {k}", callback_data=f"followup_edit:{k}"),
                         InlineKeyboardButton(f"View {k}", callback_data=f"followup_view:{k}")])
    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="admin_dashboard_menu")])
    await query.edit_message_text("🕒 *Follow-Up Rules*\n\nManage follow-up sequences per event:", reply_markup=InlineKeyboardMarkup(keyboard), parse_mode="Markdown")


async def cmd_followup_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, key = query.data.split(':', 1)
    from src.utils.followup_manager import get_followups
    seq = get_followups(key)
    if not seq:
        await query.edit_message_text(f"No follow-ups defined for *{key}*.", parse_mode="Markdown")
        return
    text = f"*Follow-ups for {key}*:\n\n"
    for i, step in enumerate(seq):
        text += f"{i+1}. Delay: {step.get('delay_hours')}h → Template: {step.get('template')}\n"
    await query.edit_message_text(text, parse_mode="Markdown")


async def cmd_followup_edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, key = query.data.split(':', 1)
    context.user_data['editing_followup_event'] = key
    # initialize buffer for line-by-line entry
    context.user_data['followup_buffer'] = []
    await query.edit_message_text(
        f"Editing follow-ups for *{key}*\n\nEnter steps one-per-line in the format:\n`<delay_hours> <TEMPLATE_KEY>`\nExample: `24 PAYMENT_REMINDER_1`\nSend each step on a new line.\nWhen finished send `DONE`. Send /cancel to abort.",
        parse_mode="Markdown"
    )
    return ENTER_TEMPLATE_TEXT


async def handle_new_followup_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin_id = update.effective_user.id
    key = context.user_data.get('editing_followup_event')
    if not key:
        await update.message.reply_text("No event selected. Use the Follow-Up Rules menu.")
        return SELECT_TEMPLATE

    text = (update.message.text or '').strip()
    # cancel handling
    if text.lower() == '/cancel':
        context.user_data.pop('editing_followup_event', None)
        context.user_data.pop('followup_buffer', None)
        await update.message.reply_text("Cancelled follow-up edit.")
        return SELECT_TEMPLATE

    # finish and save
    if text.upper() == 'DONE':
        buf = context.user_data.get('followup_buffer', [])
        try:
            seq = []
            for line in buf:
                if not line.strip():
                    continue
                # allow comma or space separated: "24, PAYMENT_REMINDER_1" or "24 PAYMENT_REMINDER_1"
                parts = [p.strip() for p in line.replace(',', ' ').split() if p.strip()]
                if len(parts) < 2:
                    raise ValueError(f"Invalid step line: '{line}'")
                delay_hours = int(parts[0])
                template = parts[1]
                seq.append({'delay_hours': delay_hours, 'template': template})

            from src.utils.followup_manager import save_followups
            save_followups(admin_id, key, seq)
            await update.message.reply_text(f"✅ Follow-up sequence for {key} saved ({len(seq)} steps).")
        except Exception as e:
            await update.message.reply_text(f"❌ Error saving follow-ups: {e}")
        finally:
            context.user_data.pop('editing_followup_event', None)
            context.user_data.pop('followup_buffer', None)
        return SELECT_TEMPLATE

    # otherwise treat line as a step and append to buffer
    buf = context.user_data.setdefault('followup_buffer', [])
    buf.append(text)
    await update.message.reply_text("Added step: {}\nSend more steps or 'DONE' to finish. Send /cancel to abort.".format(text))
    return ENTER_TEMPLATE_TEXT


async def cmd_audit_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return
    await query.answer()
    from src.utils.message_templates import tail_audit
    rows = tail_audit(limit=25)
    if not rows:
        await query.edit_message_text("No audit entries found.")
        return
    text = "*Audit (latest edits)*\n\n"
    for r in rows[-10:]:
        ts = r.get('timestamp')
        admin = r.get('admin_id')
        ek = r.get('event_key')
        text += f"{ts} — Admin {admin} — {ek}\n"
    await query.edit_message_text(text, parse_mode="Markdown")


async def cmd_admin_toggle_editing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    from src.config import SUPER_ADMIN_USER_ID
    # only super admin may toggle
    if str(query.from_user.id) != str(SUPER_ADMIN_USER_ID):
        await query.answer("Only super-admin can toggle this.", show_alert=True)
        return
    from src.utils.followup_manager import is_admin_editing_enabled, set_admin_editing_enabled
    enabled = is_admin_editing_enabled()
    set_admin_editing_enabled(not enabled)
    await query.answer()
    await query.edit_message_text(f"Admin editing enabled: {not enabled}")


async def cmd_template_preview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, key = query.data.split(':', 1)
    tpl = get_template(key)
    text = tpl.get('text') if tpl else event_registry.DEFAULT_TEMPLATES.get(key, '')
    enabled = tpl.get('enabled', True) if tpl else True
    await query.edit_message_text(f"*{key}* (enabled={enabled})\n\n{text}", parse_mode="Markdown")


async def cmd_template_edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    _, key = query.data.split(':', 1)
    context.user_data['editing_event'] = key
    tpl = get_template(key)
    text = tpl.get('text') if tpl else event_registry.DEFAULT_TEMPLATES.get(key, '')
    await query.edit_message_text(
        f"Editing *{key}*\n\nCurrent template:\n\n{text}\n\nSend the new template text using placeholders like {{name}}, or /cancel to abort.",
        parse_mode="Markdown"
    )
    return ENTER_TEMPLATE_TEXT


async def handle_new_template_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    admin_id = update.effective_user.id
    key = context.user_data.get('editing_event')
    if not key:
        await update.message.reply_text("No event selected. Use the Templates menu.")
        return SELECT_TEMPLATE
    new_text = update.message.text
    try:
        save_template(admin_id, key, new_text=new_text)
        await update.message.reply_text(f"✅ Template for {key} saved.")
    except Exception as e:
        await update.message.reply_text(f"❌ Error saving template: {e}")
    finally:
        context.user_data.pop('editing_event', None)
    return SELECT_TEMPLATE


def get_template_conversation_handler():
    from telegram.ext import ConversationHandler, MessageHandler, filters, CallbackQueryHandler

    return ConversationHandler(
        entry_points=[CallbackQueryHandler(cmd_templates_menu, pattern="^admin_templates$")],
        states={
            SELECT_TEMPLATE: [
                CallbackQueryHandler(cmd_template_preview, pattern=r"^template_preview:"),
                CallbackQueryHandler(cmd_template_edit_start, pattern=r"^template_edit:"),
            ],
            ENTER_TEMPLATE_TEXT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_new_template_text)
            ]
        },
        fallbacks=[CallbackQueryHandler(callback_back_to_admin_panel, pattern="^admin_dashboard_menu$")],
        conversation_timeout=600,  # 10 minutes timeout
        per_message=False
    )


def get_followup_conversation_handler():
    from telegram.ext import ConversationHandler, MessageHandler, filters, CallbackQueryHandler

    return ConversationHandler(
        entry_points=[CallbackQueryHandler(cmd_followup_menu, pattern="^admin_followups$")],
        states={
            SELECT_TEMPLATE: [
                CallbackQueryHandler(cmd_followup_view, pattern=r"^followup_view:"),
                CallbackQueryHandler(cmd_followup_edit_start, pattern=r"^followup_edit:"),
            ],
            ENTER_TEMPLATE_TEXT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_new_followup_text)
            ]
        },
        fallbacks=[CallbackQueryHandler(callback_back_to_admin_panel, pattern="^admin_dashboard_menu$")],
        conversation_timeout=600,  # 10 minutes timeout
        per_message=False
    )


async def cmd_member_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Display list of all members with pagination"""
    query = update.callback_query
    
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return
    
    await query.answer()
    
    # Get page number from callback data or default to 1
    try:
        page = int(query.data.split('_')[-1]) if '_' in query.data else 1
    except Exception:
        page = 1

    filter_type = _normalize_filter(context.user_data.get('members_filter', 'all'))
    per_page = MANAGE_USERS_PER_PAGE

    total_count = get_manage_users_count(filter_type, MANAGE_USERS_INACTIVE_DAYS)
    if total_count == 0:
        await query.edit_message_text("📭 No members found for this filter.")
        return

    total_pages = (total_count + per_page - 1) // per_page
    page = max(1, min(page, total_pages))

    offset = (page - 1) * per_page
    page_users = get_manage_users_page(
        filter_type=filter_type,
        limit=per_page,
        offset=offset,
        inactive_days=MANAGE_USERS_INACTIVE_DAYS
    ) or []

    # Build member list message (Name + Telegram ID only)
    message = "👥 *Member List*\n\n"
    message += f"Filter: *{_filter_label(filter_type)}*\n"
    message += f"Page {page}/{total_pages} (Total: {total_count})\n"
    message += "─────────────────────────\n"
    message += "Tap to copy the ID:\n\n"

    for user in page_users:
        name = user.get('full_name') or "Unknown"
        uid = user.get('user_id') or "?"
        message += f"• *{name}* — `{uid}`\n"
    
    # Build navigation and action buttons
    keyboard = []
    
    # Pagination buttons
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Previous", callback_data=f"admin_members_list_{page-1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("➡️ Next", callback_data=f"admin_members_list_{page+1}"))
    
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Action buttons
    keyboard.append([
        InlineKeyboardButton("✅ Paid", callback_data="admin_members_filter_paid"),
        InlineKeyboardButton("❌ Unpaid", callback_data="admin_members_filter_unpaid")
    ])
    keyboard.append([
        InlineKeyboardButton("🟢 Active", callback_data="admin_members_filter_active"),
        InlineKeyboardButton("⚪ Inactive 90d", callback_data="admin_members_filter_inactive")
    ])
    keyboard.append([
        InlineKeyboardButton("👥 All", callback_data="admin_members_filter_all"),
        InlineKeyboardButton("📥 Export Excel", callback_data=f"admin_export_excel_filter_{filter_type}")
    ])
    keyboard.append([
        InlineKeyboardButton("👤 Manage User (Enter ID)", callback_data="admin_manage_users")
    ])
    
    keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="admin_dashboard_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        text=message,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


async def callback_member_filter(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle filter selection for member list."""
    query = update.callback_query
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return

    await query.answer()

    filter_type = query.data.replace("admin_members_filter_", "", 1)
    filter_type = _normalize_filter(filter_type)
    context.user_data['members_filter'] = filter_type

    # Reset to page 1 after filter change
    await cmd_member_list(update, context)


async def cmd_manage_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Select user for management operations - HIGHEST PRIORITY state"""
    query = update.callback_query

    # Always answer callback immediately to avoid Telegram spinner
    if query:
        await query.answer()
    
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return

    # CRITICAL: GLOBAL STATE RESET - Clear ALL active conversation states
    clear_stale_states(update, context, flow_name="manage_users")
    # Explicitly drop shake/manual deduction payloads to avoid cross-talk
    context.user_data.pop('shake_deduction', None)

    # CRITICAL: Explicitly set management marker to prevent state confusion
    context.user_data["is_in_management_flow"] = True
    
    keyboard = [
        [InlineKeyboardButton("❌ Cancel", callback_data="admin_dashboard_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        text="👤 *Manage Users*\n\n"
        "Send the User ID of the member you want to manage:\n\n"
        "Example: `424837855`\n\n"
        "⚠️ Make sure to copy the exact ID (numbers only)",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    
    return MANAGE_USER_MENU


async def handle_user_id_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle user ID input for management - GUARDED state"""
    # CRITICAL: Verify we are in management flow (guard against Invoice/Store/AR interception)
    if not context.user_data.get("is_in_management_flow"):
        logger.warning(f"[MANAGE_USERS] User ID input received but not in management flow - rejecting")
        await update.message.reply_text("❌ Invalid context. Please use /menu to start over.")
        return ConversationHandler.END
    
    # CRITICAL FIX: Sanitize input with int(str().strip()) to ensure proper type handling
    # Remove ALL whitespace: leading, trailing, and internal
    input_text = str(update.message.text).strip().replace(" ", "")
    
    # Validate input is numeric BEFORE attempting int conversion
    if not input_text.isdigit():
        await update.message.reply_text(
            "❌ Invalid format. Please send a valid User ID (numbers only).\n\n"
            "Example: `424837855`\n\n"
            "💡 Tip: User IDs are numbers. If you're trying to search by name, use the member list instead.\n\n"
            "Use /cancel or click the button below to exit.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ Cancel", callback_data="admin_dashboard_menu")
            ]]),
            parse_mode="Markdown"
        )
        return MANAGE_USER_MENU
    
    try:
        # CRITICAL FIX: Use int(str().strip()) for proper type conversion (Telegram IDs are 64-bit BIGINT)
        # Double-check it's still numeric after cleaning
        if not input_text.isdigit():
            raise ValueError(f"Invalid characters in cleaned input: {input_text}")
        
        user_id = int(input_text)
        
        # Validate range (Telegram IDs are positive and reasonably large)
        if user_id <= 0:
            await update.message.reply_text(
                "❌ Invalid User ID. User IDs must be positive numbers.\n\n"
                "Example: `424837855`",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("❌ Cancel", callback_data="admin_dashboard_menu")
                ]]),
                parse_mode="Markdown"
            )
            return MANAGE_USER_MENU
            
    except ValueError as e:
        logger.error(f"[MANAGE_USERS] Failed to parse user ID '{input_text}': {e}")
        await update.message.reply_text(
            "❌ Error parsing User ID. The number might be too large or invalid.\n\n"
            "Please try again or use /cancel to exit.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ Cancel", callback_data="admin_dashboard_menu")
            ]]),
            parse_mode="Markdown"
        )
        return MANAGE_USER_MENU
    
    # CRITICAL: Ensure we're still in management flow before proceeding
    # (Prevents timeout or state confusion from breaking validation)
    if not context.user_data.get("is_in_management_flow"):
        logger.warning(f"[MANAGE_USERS] User ID {user_id} entered, but flow state lost")
        return ConversationHandler.END
    
    logger.info(f"[MANAGE_USERS] Admin {update.effective_user.id} looking up user_id={user_id} (flow confirmed)")
    
    # Get user details from database
    user = get_user(user_id)
    if not user:
        logger.warning(f"[MANAGE_USERS] User ID {user_id} not found in database")
        await update.message.reply_text(
            f"❌ User with ID `{user_id}` not found.\n\n"
            "💡 **Possible reasons:**\n"
            "• User hasn't registered yet\n"
            "• User ID was typed incorrectly\n"
            "• User was already deleted\n\n"
            "Please verify the ID and try again, or use /cancel to exit.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ Cancel", callback_data="admin_dashboard_menu")
            ]]),
            parse_mode="Markdown"
        )
        return MANAGE_USER_MENU
    
    # Store user_id in context for later use
    context.user_data['selected_user_id'] = user_id
    context.user_data['selected_user'] = user
    
    logger.info(f"[MANAGE_USERS] Found user: {user.get('full_name')} (ID: {user_id})")
    
    # Show user details and management options
    is_banned = user.get('is_banned', False)
    status = "🚫 BANNED" if is_banned else "✅ ACTIVE"
    
    message = f"👤 *User Details*\n\n"
    message += f"Name: *{user['full_name']}*\n"
    message += f"Age: {user['age']}\n"
    message += f"Gender: {user.get('gender', 'N/A')}\n"
    message += f"Role: {user['role']}\n"
    message += f"Fee Status: {user['fee_status']}\n"
    message += f"Status: {status}\n"
    message += f"Joined: {user['created_at']}\n"
    
    # Management buttons
    keyboard = [
        [InlineKeyboardButton("🚫 Ban User" if not is_banned else "✅ Unban User", 
                              callback_data="mng_usr_toggle_ban")],
        [InlineKeyboardButton("🗑️ Delete User", callback_data="mng_usr_delete")],
        [InlineKeyboardButton("🔙 Back to Menu", callback_data="admin_manage_users_back")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        message,
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    
    return SELECT_USER_ACTION


async def callback_toggle_ban(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ask for ban reason before banning user"""
    query = update.callback_query
    admin_id = query.from_user.id
    user_id = context.user_data.get('selected_user_id')
    user = context.user_data.get('selected_user')
    
    if not user_id or not user:
        await query.answer("❌ User information not found.")
        return
    
    await query.answer()
    
    is_currently_banned = user.get('is_banned', False)
    
    if is_currently_banned:
        # Unban user immediately (no flow lock needed for immediate action)
        unban_user(user_id)
        logger.info(f"Admin {admin_id} unbanned user {user_id}")
        
        # Send notification to user
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="✅ *Account Unbanned*\n\n"
                     "Your account has been unbanned and is now active. "
                     "You can access the fitness club services again.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.error(f"Could not send unban notification to user {user_id}: {e}")
        
        # Send confirmation to admin
        await query.edit_message_text(
            f"✅ User *{user['full_name']}* has been unbanned.\n\n"
            f"📢 Notification sent to user",
            parse_mode="Markdown"
        )
    else:
        # Ask for ban reason - LOCK THIS FLOW
        set_active_flow(admin_id, FLOW_BAN_USER)
        await query.edit_message_text(
            f"🚫 *Ban User: {user['full_name']}*\n\n"
            f"Please provide a reason for banning this user:\n\n"
            f"Example: \"Non-payment\", \"Disruptive behavior\", \"Suspended membership\"",
            parse_mode="Markdown"
        )
        return ENTER_BAN_REASON




async def handle_ban_reason_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle ban reason input and execute ban"""
    admin_id = update.effective_user.id
    
    # CRITICAL: Check flow ownership
    if not check_flow_ownership(admin_id, FLOW_BAN_USER):
        await update.message.reply_text("❌ Invalid context. Please use /menu to start over.")
        return ConversationHandler.END
    
    ban_reason = update.message.text.strip()
    user_id = context.user_data.get('selected_user_id')
    user = context.user_data.get('selected_user')
    admin_name = update.effective_user.first_name or "Admin"
    
    if not user_id or not user:
        await update.message.reply_text("❌ User information not found.")
        return ENTER_BAN_REASON
    
    # Ban the user with reason
    ban_user(user_id, reason=ban_reason)
    
    logger.info(f"Admin {admin_id} banned user {user_id} with reason: {ban_reason}")
    
    # Send notification to banned user
    try:
        await context.bot.send_message(
            chat_id=user_id,
            text="🚫 *Account Banned*\n\n"
                 f"Your account has been banned.\n\n"
                 f"**Reason:** {ban_reason}\n\n"
                 f"If you believe this is a mistake, please contact the fitness club administrator.",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Could not send ban notification to user {user_id}: {e}")
    
    # Send confirmation to admin
    keyboard = [
        [InlineKeyboardButton("🔙 Back to Admin Panel", callback_data="admin_dashboard_menu")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        f"✅ *Ban Executed*\n\n"
        f"User: *{user['full_name']}* (ID: `{user_id}`)\n"
        f"Reason: {ban_reason}\n\n"
        f"📢 Notification sent to user",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    
    # CRITICAL: Clear ban flow lock
    clear_active_flow(admin_id, FLOW_BAN_USER)
    
    return SELECT_USER_ACTION


async def callback_delete_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Confirm deletion of user"""
    query = update.callback_query
    admin_id = query.from_user.id
    user_id = context.user_data.get('selected_user_id')
    user = context.user_data.get('selected_user')
    
    if not user_id or not user:
        await query.answer("❌ User information not found.")
        return
    
    await query.answer()
    
    # CRITICAL: Lock DELETE_USER flow to prevent Invoice messages from interfering
    set_active_flow(admin_id, FLOW_DELETE_USER)
    
    keyboard = [
        [InlineKeyboardButton("✅ Yes, Delete", callback_data="mng_usr_confirm_delete"),
         InlineKeyboardButton("❌ Cancel", callback_data="mng_usr_cancel_delete")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        f"⚠️ *Are you sure you want to delete {user['full_name']}?*\n\n"
        "This will:\n"
        "• Remove user from database\n"
        "• Delete all activity logs\n"
        "• Delete all payment records\n\n"
        "This action cannot be undone!",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )
    
    return CONFIRM_DELETE


async def callback_confirm_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Confirm and execute user deletion"""
    query = update.callback_query
    admin_id = query.from_user.id
    
    # CRITICAL: Check flow ownership - only proceed if DELETE_USER is active
    if not check_flow_ownership(admin_id, FLOW_DELETE_USER):
        await query.answer("❌ Invalid context. Please use /menu to start over.", show_alert=True)
        return ConversationHandler.END
    
    user_id = context.user_data.get('selected_user_id')
    user = context.user_data.get('selected_user')
    admin_name = query.from_user.first_name or "Admin"
    
    if not user_id or not user:
        await query.answer("❌ User information not found.")
        return
    
    await query.answer()
    
    try:
        # Send notification to user BEFORE deletion
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text="🗑️ *Account Deleted*\n\n"
                     "Your account has been permanently deleted from the fitness club system. "
                     "All your data, including membership records and activity logs, have been removed.\n\n"
                     "If you wish to rejoin, you will need to register as a new member.",
                parse_mode="Markdown"
            )
        except Exception as e:
            logger.warning(f"Could not send deletion notification to user {user_id}: {e}")
        
        # Delete the user
        delete_user(user_id)
        logger.info(f"Admin {admin_id} deleted user {user_id}")
        
        # Send confirmation to admin
        keyboard = [
            [InlineKeyboardButton("🔙 Back to Admin Panel", callback_data="admin_dashboard_menu")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            f"✅ *User Deleted Successfully*\n\n"
            f"User: *{user['full_name']}* (ID: `{user_id}`)\n\n"
            f"📢 Notification sent to user\n"
            f"🗂️ All records removed from database",
            reply_markup=reply_markup,
            parse_mode="Markdown"
        )
        
        # Clear context
        context.user_data.pop('selected_user_id', None)
        context.user_data.pop('selected_user', None)
        
        # CRITICAL: Clear DELETE_USER flow lock
        clear_active_flow(admin_id, FLOW_DELETE_USER)
        
    except Exception as e:
        logger.error(f"Error deleting user: {e}")
        await query.edit_message_text(f"❌ Error deleting user: {str(e)}")
        # CRITICAL: Clear flow lock even on error
        clear_active_flow(admin_id, FLOW_DELETE_USER)


async def cmd_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Export all members to Excel file"""
    query = update.callback_query
    
    if not is_admin(query.from_user.id):
        await query.answer("❌ Admin access only.", show_alert=True)
        return
    
    await query.answer()
    
    try:
        filter_type = 'all'
        if query.data.startswith("admin_export_excel_filter_"):
            filter_type = _normalize_filter(query.data.replace("admin_export_excel_filter_", "", 1))

        total_count = get_manage_users_count(filter_type, MANAGE_USERS_INACTIVE_DAYS)
        if total_count == 0:
            await query.edit_message_text("📭 No members to export.")
            return

        users = get_manage_users_page(
            filter_type=filter_type,
            limit=total_count,
            offset=0,
            inactive_days=MANAGE_USERS_INACTIVE_DAYS
        ) or []
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Members"
        
        # Define headers (no phone numbers)
        headers = ["User ID", "Name", "Fee Status", "Last Activity", "Status"]
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data
        for row, user in enumerate(users, 2):
            worksheet.cell(row=row, column=1).value = user.get('user_id')
            worksheet.cell(row=row, column=2).value = user.get('full_name')
            worksheet.cell(row=row, column=3).value = user.get('fee_status') or 'unpaid'
            worksheet.cell(row=row, column=4).value = str(user.get('last_activity') or '')

            is_inactive = bool(user.get('is_inactive'))
            status = "Inactive (unpaid)" if is_inactive else "Active"
            worksheet.cell(row=row, column=5).value = status

            # Center align all cells
            for col in range(1, 6):
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 24
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 16
        
        # Save to bytes
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        
        # Send file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Members_Export_{filter_type}_{timestamp}.xlsx"
        
        await context.bot.send_document(
            chat_id=query.from_user.id,
            document=excel_file,
            filename=filename,
            caption=(
                f"📊 *Member Export*\n\n"
                f"Filter: {_filter_label(filter_type)}\n"
                f"Total members: {len(users)}\n"
                f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
        )
        
        logger.info(f"Admin {query.from_user.id} exported {len(users)} members to Excel (filter={filter_type})")
        
        # Show confirmation
        await query.edit_message_text(
            f"✅ *Export Complete*\n\n"
            f"Filter: {_filter_label(filter_type)}\n"
            f"File: {filename}\n"
            f"Members exported: {len(users)}\n"
            f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        await query.edit_message_text(f"❌ Error creating export: {str(e)}")


async def callback_back_to_admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Return to admin panel"""
    query = update.callback_query
    await query.answer()
    
    # CRITICAL: Clear any active conversation states when returning to admin panel
    # Prevents abandoned flows (Invoice, Store, User Search) from persisting
    if context.user_data:
        logger.info(f"[ADMIN_PANEL] Returning to dashboard, clearing states: {list(context.user_data.keys())}")
        context.user_data.clear()
    
    keyboard = [
        [InlineKeyboardButton("👥 Member List", callback_data="admin_members_list_1"),
         InlineKeyboardButton("📊 Dashboard", callback_data="admin_dashboard")],
        [InlineKeyboardButton("👤 Manage Users", callback_data="admin_manage_users"),
         InlineKeyboardButton("📥 Excel Export", callback_data="admin_export_excel")],
        [InlineKeyboardButton("💰 Revenue Stats", callback_data="dashboard_revenue"),
         InlineKeyboardButton("📈 Engagement", callback_data="dashboard_engagement")],
        [InlineKeyboardButton("✏️ Message Templates", callback_data="admin_templates")],
        [InlineKeyboardButton("🕒 Follow-Up Rules", callback_data="admin_followups")],
        [InlineKeyboardButton("🔒 Toggle Admin Editing", callback_data="admin_toggle_editing")],
        [InlineKeyboardButton("📜 View Audit Log", callback_data="admin_view_audit")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        text="🔧 *Admin Control Panel*\n\nSelect an option below:",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )


def get_manage_users_conversation_handler():
    """Get conversation handler for user management"""
    from telegram.ext import ConversationHandler, MessageHandler, filters, CallbackQueryHandler
    
    return ConversationHandler(
        entry_points=[CallbackQueryHandler(cmd_manage_users, pattern="^admin_manage_users$")],
        states={
            MANAGE_USER_MENU: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_user_id_input),
                CallbackQueryHandler(cmd_manage_users, pattern="^admin_manage_users$"),
            ],
            SELECT_USER_ACTION: [
                CallbackQueryHandler(callback_toggle_ban, pattern="^mng_usr_toggle_ban$"),
                CallbackQueryHandler(callback_delete_user, pattern="^mng_usr_delete$"),
                CallbackQueryHandler(cmd_manage_users, pattern="^admin_manage_users_back$"),
            ],
            CONFIRM_DELETE: [
                CallbackQueryHandler(callback_confirm_delete, pattern="^mng_usr_confirm_delete$"),
                CallbackQueryHandler(cmd_manage_users, pattern="^mng_usr_cancel_delete$"),
            ],
            ENTER_BAN_REASON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_ban_reason_input),
            ],
        },
        fallbacks=[
            CallbackQueryHandler(callback_back_to_admin_panel, pattern="^admin_dashboard_menu$"),
            CommandHandler('cancel', lambda u, c: ConversationHandler.END)
        ],
        conversation_timeout=300,  # 5 minute timeout to prevent stuck states
        allow_reentry=True,
        per_message=False,
        per_chat=True,  # CRITICAL: Isolate per chat for 200+ users
        per_user=True,  # CRITICAL: Isolate per user for admin concurrency
        name="management_flow"
    )
