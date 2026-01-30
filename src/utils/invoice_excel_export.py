"""
Invoice Report Excel Export Module
Generates Excel workbooks with invoice data and summary statistics
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_invoice_report_excel(invoices, summary, period_name):
    """
    Generate Excel workbook with invoice report
    
    Args:
        invoices: List of invoice dicts from database
        summary: Dict with summary statistics
        period_name: String describing the period (e.g., "January 2026", "Q1 2026")
        
    Returns:
        BytesIO object containing Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Add sheets
    add_summary_sheet(wb, summary, period_name)
    add_invoices_sheet(wb, invoices)
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def add_summary_sheet(wb, summary, period_name):
    """Add summary statistics sheet"""
    ws = wb.create_sheet('Summary', 0)
    
    # Header
    ws['A1'] = f'Invoice Report - {period_name}'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Empty row
    ws.row_dimensions[2].height = 5
    
    # Summary metrics
    metrics = [
        ('Total Invoices:', summary.get('total_invoices', 0)),
        ('Total Amount:', f"₹{summary.get('total_amount', 0):,.2f}"),
        ('Paid Amount:', f"₹{summary.get('paid_amount', 0):,.2f}"),
        ('Pending Amount:', f"₹{summary.get('pending_amount', 0):,.2f}"),
        ('Unique Customers:', summary.get('unique_customers', 0)),
        ('Average Invoice:', f"₹{summary.get('avg_invoice_amount', 0):,.2f}"),
    ]
    
    row = 3
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = header_fill
        
        ws[f'B{row}'] = value
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def add_invoices_sheet(wb, invoices):
    """Add detailed invoices sheet"""
    ws = wb.create_sheet('Invoices', 1)
    
    # Header row
    headers = [
        'Invoice ID',
        'Customer Name',
        'Phone',
        'Items Count',
        'Subtotal',
        'GST',
        'Shipping',
        'Total Amount',
        'Status',
        'Created Date',
        'Paid Date'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 20
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num, invoice in enumerate(invoices, 2):
        cells = [
            invoice.get('invoice_id', ''),
            invoice.get('full_name', 'N/A'),
            invoice.get('phone', ''),
            invoice.get('items_count', 0),
            invoice.get('items_subtotal', 0),
            invoice.get('gst_total', 0),
            invoice.get('shipping', 0),
            invoice.get('final_total', 0),
            invoice.get('status', 'pending').capitalize(),
            invoice.get('created_date', ''),
            invoice.get('paid_date', '') or '-'
        ]
        
        for col_num, value in enumerate(cells, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            
            # Format currency columns
            if col_num in [5, 6, 7, 8]:  # Subtotal, GST, Shipping, Total
                if isinstance(value, (int, float)):
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Center alignment for counts and dates
            elif col_num in [4, 9, 10, 11]:
                cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    column_widths = [15, 18, 15, 12, 12, 12, 12, 14, 12, 13, 13]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    return ws
