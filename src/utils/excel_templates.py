"""
Excel Template Generator for Commerce Hub
Generates sample Excel files for bulk uploads
"""

import logging
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


def generate_store_product_template() -> BytesIO:
    """
    Generate a sample Excel template for store product bulk upload.
    
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Store Products"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Product Name", "Description", "MRP", "Percentage Discount", "Final Price"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        
        # Add example rows
        examples = [
            ["Protein Powder", "Whey protein isolate 2kg", 2500, 10, 2250],
            ["Energy Bar", "High protein energy bars pack of 12", 800, 5, 760],
            ["Gym Towel", "Premium microfiber gym towel", 500, 0, 500],
        ]
        
        center_align = Alignment(horizontal="center", vertical="center")
        
        for row_idx, example in enumerate(examples, start=2):
            ws.append(example)
            for cell in ws[row_idx]:
                cell.border = border
                cell.alignment = center_align
        
        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            ["Store Product Bulk Upload Instructions"],
            [],
            ["Column Details:"],
            ["Product Name:", "Name of the product (required, max 255 characters)"],
            ["Description:", "Brief description of product (optional, max 1000 characters)"],
            ["MRP:", "Maximum Retail Price in Rs (required, numeric)"],
            ["Percentage Discount:", "Discount percentage 0-100 (optional, default 0)"],
            ["Final Price:", "Will be auto-calculated as MRP × (1 - Discount%)"],
            [],
            ["Important:"],
            ["- Do NOT modify the header row", ""],
            ["- Keep product names unique within the upload", ""],
            ["- Final Price will be calculated automatically", ""],
            ["- Use numeric values only for price and discount fields", ""],
            ["- Save file as .xlsx format before uploading", ""],
        ]
        
        for row in instructions:
            instructions_ws.append(row)
        
        # Style instructions
        instructions_ws.column_dimensions['A'].width = 30
        instructions_ws.column_dimensions['B'].width = 50
        
        title_font = Font(bold=True, size=14)
        instructions_ws['A1'].font = title_font
        
        section_font = Font(bold=True, size=11)
        for row in [3, 10]:
            instructions_ws[f'A{row}'].font = section_font
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Store product template generated successfully")
        return output
        
    except Exception as e:
        logger.error(f"Failed to generate store product template: {e}")
        return None


def generate_subscription_plan_template() -> BytesIO:
    """
    Generate a sample Excel template for subscription plan bulk upload.
    
    Returns:
        BytesIO object containing the Excel file
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscription Plans"
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Plan Name", "Duration (Days)", "Price", "Description"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 35
        
        # Add example rows
        examples = [
            ["Basic", 30, 3000, "30 days basic access"],
            ["Premium", 90, 8000, "90 days premium access"],
            ["Annual", 365, 25000, "Full year subscription"],
        ]
        
        center_align = Alignment(horizontal="center", vertical="center")
        
        for row_idx, example in enumerate(examples, start=2):
            ws.append(example)
            for cell in ws[row_idx]:
                cell.border = border
                cell.alignment = center_align
        
        # Add instructions
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            ["Subscription Plan Bulk Upload Instructions"],
            [],
            ["Column Details:"],
            ["Plan Name:", "Unique name for subscription plan (required)"],
            ["Duration (Days):", "Subscription duration in days (required, numeric)"],
            ["Price:", "Price in Rs (required, numeric)"],
            ["Description:", "Plan description (optional)"],
            [],
            ["Important:"],
            ["- Keep plan names unique", ""],
            ["- Duration must be positive number", ""],
        ]
        
        for row in instructions:
            instructions_ws.append(row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Subscription plan template generated successfully")
        return output
        
    except Exception as e:
        logger.error(f"Failed to generate subscription plan template: {e}")
        return None
