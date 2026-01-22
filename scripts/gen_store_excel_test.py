from src.utils.store_items import load_store_items
from io import BytesIO
from openpyxl import Workbook, load_workbook

items = load_store_items()
wb = Workbook()
ws = wb.active
ws.append(['Serial No','Item Name','HSN Code','MRP','GST %'])
try:
    items_sorted = sorted(items, key=lambda x: int(x.get('serial',0)))
except Exception:
    items_sorted = items
for it in items_sorted:
    ws.append([
        it.get('serial',''),
        it.get('name',''),
        it.get('hsn',''),
        float(it.get('mrp',0)) if it.get('mrp',None) is not None else '',
        float(it.get('gst',0)) if it.get('gst',None) is not None else ''
    ])

bio = BytesIO()
wb.save(bio)
bio.seek(0)

print('items_count=', len(items))
wb2 = load_workbook(filename=bio, data_only=True)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))
print('rows_read=', len(rows))
print('header=', rows[0] if rows else None)
print('sample_row_1=', rows[1] if len(rows)>1 else None)
